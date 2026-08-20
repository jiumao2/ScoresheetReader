from __future__ import annotations

import hashlib
import json
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5
from zoneinfo import ZoneInfo

from .settings import Settings

SHANGHAI = ZoneInfo("Asia/Shanghai")

# These labels are source-system join rules, not player aliases and never reach the model.
TEAM_SOURCE_LABELS: dict[tuple[str, str], str] = {
    ("男乙", "力工-先机"): "力工-先机联队",
    ("男乙", "地空-集电"): "地空+集电",
    ("男乙", "地集"): "地空+集电",
    ("男乙", "未来-现代"): "未来现代联队",
    ("男乙", "社会"): "社系",
    ("女甲", "化学"): "化院",
    ("女乙", "光经"): "光华-经济",
    ("女乙", "工学-材料"): "工材联队",
    ("女乙", "工材"): "工材联队",
    ("女乙", "数学"): "数院",
    ("女乙", "环科-哲学"): "环哲",
    ("女乙", "生历"): "生科历史联队",
    ("女乙", "生科-历史"): "生科历史联队",
    ("女乙", "社会-政管"): "社系-政管联队",
}


class MasterDataValidationError(ValueError):
    pass


@dataclass(frozen=True)
class ImportedTeam:
    id: str
    division: str
    name: str
    players: tuple[tuple[str, str], ...]


@dataclass(frozen=True)
class ImportedGame:
    id: str
    competition: str
    division: str
    date: str
    scheduled_time: str
    scheduled_at: datetime
    venue: str
    team_a_id: str | None
    team_b_id: str | None
    team_a_name: str
    team_b_name: str
    ready: bool
    unavailable_reason: str


@dataclass(frozen=True)
class MasterDataBundle:
    source_hash: str
    teams: tuple[ImportedTeam, ...]
    games: tuple[ImportedGame, ...]


def normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return " ".join(text.strip().split())


def _stable_id(kind: str, *parts: str) -> str:
    return str(uuid5(NAMESPACE_URL, ":".join(("scoresheet-reader", kind, *parts))))


def _hash_files(paths: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in paths:
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())
    return digest.hexdigest()


def _load_rosters(paths: list[Path]) -> dict[tuple[str, str], ImportedTeam]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - deployment dependency.
        raise MasterDataValidationError("读取报名表需要安装 openpyxl。") from error

    teams: dict[tuple[str, str], ImportedTeam] = {}
    collected: dict[tuple[str, str], list[str]] = {}
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for division in ("男甲", "男乙", "女甲", "女乙"):
                if division not in workbook.sheetnames:
                    continue
                for row in workbook[division].iter_rows(min_row=2, values_only=True):
                    team_name = normalize_name(row[0] if len(row) > 0 else "")
                    player_name = normalize_name(row[1] if len(row) > 1 else "")
                    if not team_name or not player_name:
                        continue
                    collected.setdefault((division, team_name), []).append(player_name)
        finally:
            workbook.close()

    for (division, team_name), names in collected.items():
        duplicates = sorted({name for name in names if names.count(name) > 1})
        if duplicates:
            raise MasterDataValidationError(
                f"{division} {team_name} 存在重复唯一姓名：{', '.join(duplicates)}。"
            )
        team_id = _stable_id("team", division, team_name)
        players = tuple((_stable_id("player", team_id, name), name) for name in names)
        teams[(division, team_name)] = ImportedTeam(
            id=team_id,
            division=division,
            name=team_name,
            players=players,
        )
    return teams


def _schedule_team(
    division: str,
    source_name: object,
    teams: dict[tuple[str, str], ImportedTeam],
) -> tuple[ImportedTeam | None, str]:
    raw = normalize_name(source_name)
    canonical = TEAM_SOURCE_LABELS.get((division, raw), raw)
    return teams.get((division, canonical)), canonical


def _load_schedule(
    path: Path,
    competition: str,
    teams: dict[tuple[str, str], ImportedTeam],
) -> list[ImportedGame]:
    games: list[ImportedGame] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError as error:
            raise MasterDataValidationError(f"赛程第 {line_number} 行不是有效 JSON。") from error
        division = normalize_name(row.get("group"))
        team_a, team_a_name = _schedule_team(division, row.get("home_team"), teams)
        team_b, team_b_name = _schedule_team(division, row.get("away_team"), teams)
        unresolved: list[str] = []
        if team_a is None:
            unresolved.append(team_a_name)
        if team_b is None:
            unresolved.append(team_b_name)
        raw_time = str((row.get("time") or {}).get("$date") or "")
        try:
            scheduled_at = datetime.fromisoformat(raw_time.replace("Z", "+00:00")).astimezone(
                SHANGHAI
            )
        except ValueError as error:
            raise MasterDataValidationError(
                f"比赛 {row.get('_id', line_number)} 的时间格式无效。"
            ) from error
        games.append(
            ImportedGame(
                id=str(row.get("_id") or _stable_id("game", str(line_number))),
                competition=competition,
                division=division,
                date=scheduled_at.strftime("%Y-%m-%d"),
                scheduled_time=scheduled_at.strftime("%H:%M"),
                scheduled_at=scheduled_at,
                venue=normalize_name(row.get("place")),
                team_a_id=team_a.id if team_a else None,
                team_b_id=team_b.id if team_b else None,
                team_a_name=team_a.name if team_a else team_a_name,
                team_b_name=team_b.name if team_b else team_b_name,
                ready=not unresolved,
                unavailable_reason=(
                    "等待确定参赛球队：" + "、".join(unresolved) if unresolved else ""
                ),
            )
        )
    return games


def _load_fixture(path: Path) -> MasterDataBundle:
    payload = json.loads(path.read_text(encoding="utf-8"))
    teams: list[ImportedTeam] = []
    team_index: dict[str, ImportedTeam] = {}
    for item in payload.get("teams", []):
        division = normalize_name(item["division"])
        name = normalize_name(item["name"])
        names = [normalize_name(value) for value in item.get("player_names", [])]
        if len(names) != len(set(names)):
            raise MasterDataValidationError(f"测试主数据 {division} {name} 存在重复姓名。")
        team_id = _stable_id("team", division, name)
        team = ImportedTeam(
            id=team_id,
            division=division,
            name=name,
            players=tuple((_stable_id("player", team_id, player), player) for player in names),
        )
        teams.append(team)
        team_index[f"{division}:{name}"] = team
    games: list[ImportedGame] = []
    for item in payload.get("games", []):
        division = normalize_name(item["division"])
        team_a_name = normalize_name(item["team_a"])
        team_b_name = normalize_name(item["team_b"])
        team_a = team_index.get(f"{division}:{team_a_name}")
        team_b = team_index.get(f"{division}:{team_b_name}")
        scheduled_at = datetime.fromisoformat(f"{item['date']}T{item['scheduled_time']}:00+08:00")
        games.append(
            ImportedGame(
                id=str(item["id"]),
                competition=normalize_name(item.get("competition", "合成测试赛")),
                division=division,
                date=str(item["date"]),
                scheduled_time=str(item["scheduled_time"]),
                scheduled_at=scheduled_at,
                venue=normalize_name(item.get("venue")),
                team_a_id=team_a.id if team_a else None,
                team_b_id=team_b.id if team_b else None,
                team_a_name=team_a.name if team_a else team_a_name,
                team_b_name=team_b.name if team_b else team_b_name,
                ready=team_a is not None and team_b is not None,
                unavailable_reason="" if team_a and team_b else "测试主数据球队未解析。",
            )
        )
    return MasterDataBundle(
        source_hash=_hash_files([path]),
        teams=tuple(teams),
        games=tuple(games),
    )


def load_master_data(settings: Settings) -> MasterDataBundle | None:
    fixture = settings.resolved_master_fixture_path
    if fixture is not None:
        return _load_fixture(fixture) if fixture.exists() else None

    directory = settings.resolved_master_data_dir
    schedule_candidates = sorted(directory.glob("Schedule_*.json")) if directory.exists() else []
    male = directory / "男篮.xlsx"
    female = directory / "女篮.xlsx"
    if not schedule_candidates or not male.exists() or not female.exists():
        return None
    files = [schedule_candidates[0], male, female]
    roster = _load_rosters([male, female])
    return MasterDataBundle(
        source_hash=_hash_files(files),
        teams=tuple(roster.values()),
        games=tuple(_load_schedule(schedule_candidates[0], settings.competition_name, roster)),
    )
